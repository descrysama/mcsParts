from selenium import webdriver;
from selenium.webdriver.common.by import By;
from selenium.webdriver.support.ui import WebDriverWait;
from selenium.webdriver.support import expected_conditions as EC;
import openpyxl;
import os;

def utopya_single(driver: webdriver, url):
    items = []
    no_sku_url = []

    get_sku(items, no_sku_url, url, driver)

        
    
    if(len(no_sku_url) > 0) :
        for url in no_sku_url :
            item = ['', '']
            driver.get(url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "li.attr-sku")))
            li = driver.find_element(by="css selector", value="li.attr-sku")
            sku = li.find_element(by="tag name", value="strong").text
            item[0] = sku.strip()

            full_price_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.price-box span.price")))
            full_price = full_price_element.text.replace("€", ",")
            full_price_check = full_price.split(",")
            if not full_price_check[1]:
                full_price = full_price_check[0]
            item[1] = full_price
            items.append(item)     

    file_path = './utopya/src/output.xlsx'
    file_exists = os.path.isfile(file_path)
    if file_exists:
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    worksheet = workbook.active
    last_row = worksheet.max_row + 1
    for row_number, row in enumerate(items, start=last_row):
        worksheet.cell(row=row_number, column=1, value=row[0])
        worksheet.cell(row=row_number, column=2, value=row[1])
    workbook.save(file_path)

    return


def get_sku(items, no_sku_url, url, driver):

    driver.get(url)
    find_element_with_retry(items, no_sku_url, url, driver)
    products = driver.find_elements(by="css selector", value="div.product.details.product-item-details")

    for product in products:
            ##Recuperer le sku et si il n'y en a pas push l'url dans le tableau
        try :
            item = ['', '']
            form = product.find_element(by="tag name", value="form")
            item[0] = form.get_attribute("data-product-sku")
            ## Recuperer le prix
            pricing = product.find_element(by="css selector", value="div.pricing")
            price_div = pricing.find_element(by="css selector", value="span.price.price-final_price.tax.weee")
            item[1] = price_div.text.replace("€", ",")
            items.append(item)
        except: 
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "a.product-item-link.name")))
                url = product.find_element(by="css selector", value="a.product-item-link.name")
                no_sku_url.append(url.get_attribute("href"))
            except:
                get_sku()


def find_element_with_retry(items, no_sku_url, url, driver):
    while True:
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.pricing")))
            return
        except:
            get_sku(items, no_sku_url, url, driver)
            continue